# -*- coding: utf-8 -*-
"""
适合同一文件，已经存在并且经常会读取写入的，速度较快，大文件也可以，功能一般
"""
from openpyxl import load_workbook
from openpyxl import Workbook
import datetime

# 不支持xls格式
wb2 = load_workbook('/Users/whf/dev/code_practice/python_models/execl/(明细).xlsx')

wb = Workbook()     # 实例化
ws = wb.active      # 激活 worksheet
ws['A1'] = 42        # 方式一：数据可以直接分配到单元格中(可以输入公式)
ws.append([1, 2, 3])        # 方式二：可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)
ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")  # 方式三：Python 类型会被自动转换

ws1 = wb.create_sheet("Mysheet")    # 方式一：插入到最后(default)
ws1 = wb.create_sheet("Mysheet", 0)   # 方式二：插入到最开始的位置

ws3 = wb2["出货单明细信息"]    # sheet 名称可以作为 key 进行索引
ws4 = wb2.get_sheet_by_name("出货单明细信息")
print(wb2.sheetnames)
